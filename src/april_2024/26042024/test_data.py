import time
from openpyxl import load_workbook
import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (ElementNotVisibleException, ElementNotSelectableException)
import os


# Data Driven Test Case for the Login Page.
# Invalid Login for the VWO Page

def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row, values_only=True):
        username, password = row
        credentials.append({"username": username, "password": password})
        return credentials


file_path_fromos = os.getcwd() + " /py2xtestdata.xlsx "
print(file_path_fromos)


@pytest.mark.parametrize("user cred", read_credentials_from_excel(file_path_fromos))
# @pytest.fixture()
@allure.title("Verify the invalid Login with the Excel Testdata")
@allure.description("TC #1 -10 invalid Login verification for app.vwo.com")
def test_vwo_login(user_cred):
    username = user_cred["username"]
    password = user_cred["password"]
    # vwo_login(username=username, password=password)
    print(username, password)


def vwo_login(username, password):  # * made this as parametrized so that i can run multiple times
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://app.vwo.com/")

    email_address_ele = driver.find_element(By.CSS_SELECTOR, "[name='username']")
    email_address_ele.send_keys(username)
    # email_address_ele.send_keys("contact+atb5x@thetestingacademy.com")

    password_ele = driver.find_element(By.CSS_SELECTOR, "[name='password']")
    password_ele.send_keys(password)
    # password_ele.send_keys("Wingify@SDET")

    button_submit_element = driver.find_element(By.ID, "js-login-btn")
    button_submit_element.click()

    time.sleep(5)
    result = driver.current_url
    print(result)

    if result !="https://app.vwo.com/#/dashboard":
        print("Invalid Login")
    else:
        print("Valid Login")
    driver.quit()